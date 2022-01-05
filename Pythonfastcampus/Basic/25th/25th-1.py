from openpyxl import load_workbook

wb = load_workbook('test_data.xlsx', read_only=True) # read_only 읽기전용으로 바꿈 / 행을 순서대로 읽기때문에 원하는 열을 출력하는 건 불가능 
data = wb.active

for row in data.iter_rows(): # 데이터를 행단위로 가져옴 / max_col, max_row 사용
    for cell in row:            
        print(cell.value)